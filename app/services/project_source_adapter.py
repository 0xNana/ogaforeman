"""Structured text ingestion boundary for project initialization sources."""

from __future__ import annotations

import csv
import re
import unicodedata
from collections.abc import Iterable
from datetime import date, datetime
from dataclasses import dataclass, field
from hashlib import sha256
from io import BytesIO, StringIO
from zipfile import BadZipFile, ZipFile

import xlrd
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from app.domain.project_import import SourceType


class StructuredTextInputError(ValueError):
    code = "PROJECT_SOURCE_INVALID"


@dataclass(frozen=True, slots=True)
class StructuredTextSource:
    """Normalized source payload handed to the extraction workflow."""

    name: str
    source_type: SourceType
    text: str
    checksum: str


@dataclass(slots=True)
class _BoundedTextLines:
    lines: list[str] = field(default_factory=list)
    byte_size: int = 0

    def append_line(self, value: str) -> None:
        stripped = value.strip()
        if not stripped:
            return
        self.byte_size += len(stripped.encode("utf-8")) + 1
        if self.byte_size > 800_000:
            raise StructuredTextInputError("extracted project text exceeds the input limit")
        self.lines.append(stripped)

    def append_row(self, values: Iterable[object]) -> None:
        cells = [self._cell_text(value) for value in values]
        while cells and not cells[-1]:
            cells.pop()
        if any(cells):
            self.append_line(" | ".join(cells))

    def require_text(self) -> str:
        if not self.lines:
            raise StructuredTextInputError("project file has no extractable text")
        return "\n".join(self.lines)

    @staticmethod
    def _cell_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return str(value).strip()


class StructuredTextProjectAdapter:
    """Accept reasonable pasted text, Markdown, and OG-template variation.

    This adapter deliberately does not extract canonical entities. It only
    normalizes source text so the ADK/Gemini extraction boundary receives a
    stable, bounded document while preserving evidence and unresolved dates.
    """

    # Keep inline Firestore source documents safely below the 1 MiB document
    # limit, including UTF-8 expansion and document metadata.
    _MAX_SOURCE_CHARS = 800_000
    _DATE_LABELS = re.compile(r"^(due|date|finish|finished by|planned finish)\s*:\s*(.+)$", re.I)
    _TASK_LABEL = re.compile(r"^(?:task|activity|work item)\s*:\s*(.+)$", re.I)
    _DEPENDENCY_LABEL = re.compile(r"^(?:depends on|dependency|predecessor)\s*:\s*(.+)$", re.I)
    _MATERIALS_HEADING = re.compile(r"^(?:materials?|material requirements?)\s*:?\s*$", re.I)

    def __init__(
        self, *, name: str = "pasted-project.txt", source_type: SourceType | None = None
    ) -> None:
        if not name.strip():
            raise StructuredTextInputError("source name cannot be empty")
        self._name = name.strip()
        self._source_type = source_type or self._infer_source_type(name)

    def load(self, text: str) -> StructuredTextSource:
        normalized = self.normalize_input(text)
        return StructuredTextSource(
            name=self._name,
            source_type=self._source_type,
            text=normalized,
            checksum=sha256(normalized.encode("utf-8")).hexdigest(),
        )

    def normalize_input(self, text: str) -> str:
        if not isinstance(text, str):
            raise StructuredTextInputError("structured project source must be text")
        if not text.strip():
            raise StructuredTextInputError("structured project source cannot be empty")
        if len(text) > self._MAX_SOURCE_CHARS or len(text.encode("utf-8")) > 800_000:
            raise StructuredTextInputError("structured project source exceeds the input limit")

        normalized_lines: list[str] = []
        for raw_line in (
            unicodedata.normalize("NFKC", text)
            .replace("\r\n", "\n")
            .replace("\r", "\n")
            .split("\n")
        ):
            line = raw_line.strip()
            if not line:
                if normalized_lines and normalized_lines[-1] != "":
                    normalized_lines.append("")
                continue
            line = re.sub(r"^(?:[-*+]\s+|\d+[.)]\s+)", "", line)
            line = self._normalize_label(line)
            normalized_lines.append(line)

        while normalized_lines and normalized_lines[-1] == "":
            normalized_lines.pop()
        if not normalized_lines:
            raise StructuredTextInputError("structured project source cannot be empty")
        return "\n".join(normalized_lines) + "\n"

    def extract(self, text: str) -> str:
        """Return normalized source text for the ADK/Gemini extraction step."""

        return self.load(text).text

    def _normalize_label(self, line: str) -> str:
        heading = re.match(r"^#{1,6}\s+(.+)$", line)
        if heading:
            line = heading.group(1).strip()
        task = self._TASK_LABEL.match(line)
        if task:
            return f"Task: {task.group(1).strip()}"
        dependency = self._DEPENDENCY_LABEL.match(line)
        if dependency:
            return f"Depends on: {dependency.group(1).strip()}"
        date_match = self._DATE_LABELS.match(line)
        if date_match:
            return f"Due: {date_match.group(2).strip()}"
        if self._MATERIALS_HEADING.match(line):
            return "Materials:"
        return line

    @staticmethod
    def _infer_source_type(name: str) -> SourceType:
        lowered = name.casefold()
        if lowered.endswith((".md", ".markdown")):
            return SourceType.MARKDOWN
        return SourceType.TEXT


class ProjectDocumentAdapter:
    """Convert supported office documents into bounded structured text."""

    MAX_FILE_BYTES = 10_000_000
    _MAX_ARCHIVE_BYTES = 50_000_000
    _MAX_ARCHIVE_ENTRIES = 10_000
    _MAX_ROWS = 20_000
    _MAX_CELLS = 200_000
    _MAX_PDF_PAGES = 500
    _MAX_PDF_PAGE_CONTENT_BYTES = 10_000_000
    _EXPECTED_TYPES = {
        ".docx": SourceType.FILE,
        ".pdf": SourceType.FILE,
        ".xlsx": SourceType.SPREADSHEET,
        ".xls": SourceType.SPREADSHEET,
        ".csv": SourceType.SPREADSHEET,
    }

    def __init__(self, *, name: str, source_type: SourceType) -> None:
        normalized_name = name.strip()
        if not normalized_name:
            raise StructuredTextInputError("source name cannot be empty")
        extension = self._extension(normalized_name)
        expected_type = self._EXPECTED_TYPES.get(extension)
        if expected_type is None or source_type is not expected_type:
            raise StructuredTextInputError("project file type does not match its extension")
        self._name = normalized_name
        self._extension_value = extension
        self._source_type = source_type

    def load(self, content: bytes) -> StructuredTextSource:
        if not isinstance(content, bytes):
            raise StructuredTextInputError("project file content must be bytes")
        if not content:
            raise StructuredTextInputError("project file cannot be empty")
        if len(content) > self.MAX_FILE_BYTES:
            raise StructuredTextInputError("project file exceeds the input limit")

        try:
            extracted = self._extract(content)
        except StructuredTextInputError:
            raise
        except Exception as exc:
            raise StructuredTextInputError("project file could not be read") from exc
        return StructuredTextProjectAdapter(
            name=self._name,
            source_type=self._source_type,
        ).load(extracted)

    def _extract(self, content: bytes) -> str:
        if self._extension_value == ".docx":
            return self._extract_docx(content)
        if self._extension_value == ".pdf":
            return self._extract_pdf(content)
        if self._extension_value == ".xlsx":
            return self._extract_xlsx(content)
        if self._extension_value == ".xls":
            return self._extract_xls(content)
        return self._extract_csv(content)

    def _extract_docx(self, content: bytes) -> str:
        self._validate_archive(content)
        document = Document(BytesIO(content))
        lines = _BoundedTextLines()
        for block in document.iter_inner_content():
            if isinstance(block, Paragraph):
                lines.append_line(block.text)
            elif isinstance(block, Table):
                for row in block.rows:
                    lines.append_row(cell.text for cell in row.cells)
        return lines.require_text()

    def _extract_pdf(self, content: bytes) -> str:
        if not content.startswith(b"%PDF-"):
            raise StructuredTextInputError("PDF signature is invalid")
        reader = PdfReader(BytesIO(content), strict=False)
        if reader.is_encrypted and reader.decrypt("") == 0:
            raise StructuredTextInputError("password-protected PDFs are not supported")
        if len(reader.pages) > self._MAX_PDF_PAGES:
            raise StructuredTextInputError("PDF exceeds the page limit")
        lines = _BoundedTextLines()
        for page_number, page in enumerate(reader.pages, start=1):
            page_content = page.get_contents()
            if page_content is None:
                continue
            if len(page_content.get_data()) > self._MAX_PDF_PAGE_CONTENT_BYTES:
                raise StructuredTextInputError("PDF page content exceeds the input limit")
            page_text = page.extract_text(extraction_mode="layout") or ""
            if page_text.strip():
                lines.append_line(f"[Page {page_number}]")
                for line in page_text.splitlines():
                    lines.append_line(line)
        if not lines.lines:
            raise StructuredTextInputError("PDF has no extractable text; scanned PDFs require OCR")
        return lines.require_text()

    def _extract_xlsx(self, content: bytes) -> str:
        self._validate_archive(content)
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        lines = _BoundedTextLines()
        row_count = 0
        cell_count = 0
        try:
            for worksheet in workbook.worksheets:
                lines.append_line(f"[Sheet: {worksheet.title}]")
                for row in worksheet.iter_rows(values_only=True):
                    row_count += 1
                    cell_count += len(row)
                    self._enforce_spreadsheet_limits(row_count, cell_count)
                    lines.append_row(row)
        finally:
            workbook.close()
        return lines.require_text()

    def _extract_xls(self, content: bytes) -> str:
        if not content.startswith(bytes.fromhex("D0CF11E0A1B11AE1")):
            raise StructuredTextInputError("legacy Excel signature is invalid")
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        lines = _BoundedTextLines()
        row_count = 0
        cell_count = 0
        try:
            for worksheet in workbook.sheets():
                lines.append_line(f"[Sheet: {worksheet.name}]")
                for row_index in range(worksheet.nrows):
                    row_count += 1
                    cell_count += worksheet.ncols
                    self._enforce_spreadsheet_limits(row_count, cell_count)
                    lines.append_row(worksheet.row_values(row_index))
        finally:
            workbook.release_resources()
        return lines.require_text()

    def _extract_csv(self, content: bytes) -> str:
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = content.decode("cp1252")
        disallowed_controls = sum(
            character < " " and character not in "\t\n\r" for character in decoded
        )
        if "\x00" in decoded or disallowed_controls > max(1, len(decoded) // 100):
            raise StructuredTextInputError("CSV contains unsupported binary content")
        lines = _BoundedTextLines()
        cell_count = 0
        for row_count, row in enumerate(csv.reader(StringIO(decoded)), start=1):
            cell_count += len(row)
            self._enforce_spreadsheet_limits(row_count, cell_count)
            lines.append_row(row)
        return lines.require_text()

    def _validate_archive(self, content: bytes) -> None:
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
                if len(entries) > self._MAX_ARCHIVE_ENTRIES:
                    raise StructuredTextInputError("project file archive has too many entries")
                if sum(entry.file_size for entry in entries) > self._MAX_ARCHIVE_BYTES:
                    raise StructuredTextInputError("project file archive expands beyond the limit")
        except BadZipFile as exc:
            raise StructuredTextInputError("project file archive is invalid") from exc

    @staticmethod
    def _extension(name: str) -> str:
        position = name.rfind(".")
        return name[position:].casefold() if position >= 0 else ""

    @staticmethod
    def _enforce_spreadsheet_limits(row_count: int, cell_count: int) -> None:
        if row_count > ProjectDocumentAdapter._MAX_ROWS:
            raise StructuredTextInputError("spreadsheet exceeds the row limit")
        if cell_count > ProjectDocumentAdapter._MAX_CELLS:
            raise StructuredTextInputError("spreadsheet exceeds the cell limit")


__all__ = [
    "ProjectDocumentAdapter",
    "StructuredTextInputError",
    "StructuredTextProjectAdapter",
    "StructuredTextSource",
]
